"""
scripts/prepare_tulip.py — Convert the TULIP dataset into pipeline-compatible CSVs.

The TULIP dataset (TU-Leuven Integrated Parkinsonism) uses a different
directory layout and labelling scheme from the primary PD-PS dataset.
This script reads the TULIP directory tree and its Excel/CSV labels,
then writes the three CSV files that ``run_pipeline.py`` expects:

  1. **vid_score.csv**   — one row per (subject, hand, camera) video
                           with a ``video_path`` column.
  2. **scores.csv**      — MDS-UPDRS pronation-supination scores per
                           subject/hand, compatible with the orchestrator
                           merge on ``(ids, visit, medication_state, hand)``.
  3. **id2vid.csv**      — trivial identity mapping (video_id == patient_id)
                           so the existing ``load_videoid_to_patientid_map``
                           lookup works unchanged.

Usage
-----
    python scripts/prepare_tulip.py \\
        --tulip-root /path/to/tulip-dataset \\
        --output-dir /path/to/tulip-dataset/pipeline_csvs

    # Then run the pipeline with the TULIP config:
    python scripts/run_pipeline.py --config configs/config_tulip.yaml
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

PS_LEFT_FOLDER = "13. Pronation_and_supination_left"
PS_RIGHT_FOLDER = "14. Pronation_and_supination_right"

HAND_MAP = {
    PS_LEFT_FOLDER: "Left",
    PS_RIGHT_FOLDER: "Right",
}

UPDRS_PS_NAMES = {
    "Pronation-supination - Right hand": "Right",
    "Pronation-supination - Left hand": "Left",
}

# Score aggregation: median of three clinicians, rounded to nearest int.
SCORE_COLUMN = "ProS"


# ─────────────────────────────────────────────────────────────────────────────
# Discovery
# ─────────────────────────────────────────────────────────────────────────────


def discover_ps_videos(tulip_root: Path) -> list[dict]:
    """Walk the TULIP tree and collect all pronation-supination video paths.

    Returns a list of dicts with keys:
        subject_id, subject_num, hand, camera, video_path
    """
    rows: list[dict] = []
    subject_dirs = sorted(
        [d for d in tulip_root.iterdir() if d.is_dir() and d.name.startswith("Subject_")],
        key=lambda d: int(re.search(r"(\d+)", d.name).group(1)),
    )

    for subj_dir in subject_dirs:
        subj_num = int(re.search(r"(\d+)", subj_dir.name).group(1))
        subj_id = f"TULIP_{subj_num:03d}"

        for folder_name, hand in HAND_MAP.items():
            task_dir = subj_dir / folder_name
            if not task_dir.is_dir():
                continue

            mp4s = sorted(task_dir.glob("*.mp4"))
            for mp4 in mp4s:
                cam_match = re.search(r"[Cc]amera\s*(\d+)", mp4.stem)
                camera = int(cam_match.group(1)) if cam_match else 0
                rows.append(
                    {
                        "subject_id": subj_id,
                        "subject_num": subj_num,
                        "hand": hand,
                        "camera": camera,
                        "video_path": str(mp4),
                    }
                )

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Label loading
# ─────────────────────────────────────────────────────────────────────────────


def load_ps_scores_from_xlsx(xlsx_path: Path) -> pd.DataFrame:
    """Parse the master TULIP xlsx and extract PS scores.

    Returns a DataFrame with columns:
        subject_num, hand, clinician_1, clinician_2, clinician_3, ProS
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read the TULIP xlsx. "
            "Install with: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records: list[dict] = []
    for row in all_rows[2:]:  # skip 2-row header
        subj_raw, updrs_name = row[0], row[1]
        if subj_raw is None or updrs_name is None:
            continue
        if str(updrs_name) not in UPDRS_PS_NAMES:
            continue

        hand = UPDRS_PS_NAMES[str(updrs_name)]
        c1, c2, c3 = row[2], row[3], row[4]

        # Skip subjects with no scores
        scores = []
        for s in (c1, c2, c3):
            if s is not None:
                try:
                    scores.append(int(s))
                except (ValueError, TypeError):
                    pass

        if not scores:
            continue

        median_score = int(np.round(np.median(scores)))

        records.append(
            {
                "subject_num": int(subj_raw),
                "hand": hand,
                "clinician_1": c1,
                "clinician_2": c2,
                "clinician_3": c3,
                SCORE_COLUMN: median_score,
            }
        )

    return pd.DataFrame(records)


def load_ps_scores_from_csvs(labels_dir: Path) -> pd.DataFrame:
    """Fallback: load PS scores from per-subject CSVs in labels_csv_files/."""
    records: list[dict] = []

    for csv_file in sorted(labels_dir.glob("subject*_labels.csv")):
        num_match = re.search(r"subject(\d+)", csv_file.stem)
        if not num_match:
            continue
        subj_num = int(num_match.group(1))

        df = pd.read_csv(csv_file, encoding="utf-8-sig")
        for _, row in df.iterrows():
            updrs_name = str(row.get("UPDRS_name", ""))
            if updrs_name not in UPDRS_PS_NAMES:
                continue
            hand = UPDRS_PS_NAMES[updrs_name]

            scores = []
            for col in ("label_clinician1", "label_clinician2", "label_clinician3"):
                val = row.get(col)
                if val is not None and not pd.isna(val):
                    try:
                        scores.append(int(val))
                    except (ValueError, TypeError):
                        pass
            if not scores:
                continue

            median_score = int(np.round(np.median(scores)))
            records.append(
                {
                    "subject_num": subj_num,
                    "hand": hand,
                    "clinician_1": scores[0] if len(scores) > 0 else None,
                    "clinician_2": scores[1] if len(scores) > 1 else None,
                    "clinician_3": scores[2] if len(scores) > 2 else None,
                    SCORE_COLUMN: median_score,
                }
            )

    return pd.DataFrame(records)


def load_diagnoses_from_xlsx(xlsx_path: Path) -> dict[int, str]:
    """Extract overall HT/PD diagnosis per subject from the xlsx right-side table."""
    try:
        import openpyxl
    except ImportError:
        return {}

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    diagnoses: dict[int, str] = {}
    for row in all_rows[3:]:  # row 3 onward has subject data in cols 6-9
        subj_raw = row[6]
        if subj_raw is None:
            break
        try:
            subj_num = int(subj_raw)
        except (ValueError, TypeError):
            continue

        # Majority vote across 3 clinicians
        votes = [str(row[i]) for i in (7, 8, 9) if row[i] is not None]
        if not votes:
            continue
        pd_count = sum(1 for v in votes if v.upper() == "PD")
        diagnoses[subj_num] = "PD" if pd_count > len(votes) / 2 else "HT"

    return diagnoses


# ─────────────────────────────────────────────────────────────────────────────
# CSV generation
# ─────────────────────────────────────────────────────────────────────────────


def generate_csvs(
    tulip_root: Path,
    output_dir: Path,
    cameras: Optional[list[int]] = None,
) -> dict[str, Path]:
    """Generate pipeline-compatible CSVs from the TULIP dataset.

    Parameters
    ----------
    tulip_root : Path
        Root of the TULIP dataset.
    output_dir : Path
        Directory to write the generated CSVs into.
    cameras : list[int] | None
        If provided, only include these camera numbers (1-indexed).
        Default: include all cameras.

    Returns
    -------
    dict with keys ``vid_score_path``, ``score_csv_path``, ``id2vid_csv_path``
    mapping to the generated file paths.
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── 1. Discover videos ──────────────────────────────────────────────
    videos = discover_ps_videos(tulip_root)
    if not videos:
        raise RuntimeError(f"No pronation-supination videos found under {tulip_root}")

    vid_df = pd.DataFrame(videos)
    if cameras:
        vid_df = vid_df[vid_df["camera"].isin(cameras)].copy()
        if vid_df.empty:
            raise RuntimeError(f"No videos remaining after camera filter {cameras}")

    print(f"  Discovered {len(vid_df)} PS videos across "
          f"{vid_df['subject_num'].nunique()} subjects")

    # ── 2. Load scores ──────────────────────────────────────────────────
    xlsx_path = tulip_root / "Final_Label_TULIPver1.0_CVPR_2025_Mar.xlsx"
    labels_dir = tulip_root / "labels_csv_files"

    if xlsx_path.exists():
        print(f"  Loading scores from: {xlsx_path.name}")
        scores_df = load_ps_scores_from_xlsx(xlsx_path)
    elif labels_dir.is_dir():
        print(f"  Loading scores from: {labels_dir.name}/")
        scores_df = load_ps_scores_from_csvs(labels_dir)
    else:
        raise FileNotFoundError(
            f"No label source found. Expected either:\n"
            f"  {xlsx_path}\n"
            f"  {labels_dir}/"
        )

    if scores_df.empty:
        raise RuntimeError("No PS scores could be extracted from the label files")

    # Add subject_id column
    scores_df["subject_id"] = scores_df["subject_num"].apply(lambda n: f"TULIP_{n:03d}")

    print(f"  Loaded PS scores for {len(scores_df)} (subject, hand) pairs")
    print(f"  Score distribution: {dict(scores_df[SCORE_COLUMN].value_counts().sort_index())}")

    # ── 3. Load diagnoses ───────────────────────────────────────────────
    diagnoses = {}
    if xlsx_path.exists():
        diagnoses = load_diagnoses_from_xlsx(xlsx_path)
    if diagnoses:
        print(f"  Diagnoses: "
              f"{sum(1 for v in diagnoses.values() if v == 'PD')} PD, "
              f"{sum(1 for v in diagnoses.values() if v == 'HT')} HT")

    # ── 4. Write vid_score.csv ──────────────────────────────────────────
    vid_score_path = output_dir / "tulip_vid_score.csv"
    vid_out = vid_df[["video_path"]].copy()
    vid_out.to_csv(vid_score_path, index=False)
    print(f"  Written: {vid_score_path}  ({len(vid_out)} rows)")

    # ── 5. Write scores.csv ─────────────────────────────────────────────
    # Pipeline merges on: ids, visit, medication_state, hand
    score_csv_path = output_dir / "tulip_scores.csv"
    score_out = scores_df[["subject_id", "hand", SCORE_COLUMN]].copy()
    score_out = score_out.rename(columns={"subject_id": "ids"})
    score_out["visit"] = 1
    score_out["medication_state"] = "Off"
    if diagnoses:
        score_out["diagnosis"] = score_out["ids"].map(
            {f"TULIP_{k:03d}": v for k, v in diagnoses.items()}
        )
    # Reorder columns
    cols = ["ids", "visit", "medication_state", "hand", SCORE_COLUMN]
    if "diagnosis" in score_out.columns:
        cols.append("diagnosis")
    score_out = score_out[cols]
    score_out.to_csv(score_csv_path, index=False)
    print(f"  Written: {score_csv_path}  ({len(score_out)} rows)")

    # ── 6. Write id2vid.csv ─────────────────────────────────────────────
    # The pipeline uses id2vid to map video_id -> patient_id.
    # For TULIP, subject_id IS the patient_id, so this is an identity map.
    id2vid_path = output_dir / "tulip_id2vid.csv"
    unique_ids = sorted(vid_df["subject_id"].unique())
    id2vid_rows = [[sid, f"('{sid}',)"] for sid in unique_ids]
    id2vid_df = pd.DataFrame(id2vid_rows)
    id2vid_df.to_csv(id2vid_path, index=False, header=False)
    print(f"  Written: {id2vid_path}  ({len(id2vid_rows)} rows)")

    # ── 7. Summary ──────────────────────────────────────────────────────
    # Show which subjects have videos but no scores (or vice versa)
    vid_subjects = set(vid_df["subject_num"].unique())
    score_subjects = set(scores_df["subject_num"].unique())
    missing_scores = vid_subjects - score_subjects
    missing_videos = score_subjects - vid_subjects
    if missing_scores:
        print(f"\n  WARNING: subjects with videos but no scores: {sorted(missing_scores)}")
    if missing_videos:
        print(f"  WARNING: subjects with scores but no videos: {sorted(missing_videos)}")

    return {
        "vid_score_path": vid_score_path,
        "score_csv_path": score_csv_path,
        "id2vid_csv_path": id2vid_path,
    }


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Prepare TULIP dataset for the PD-PS kinematic pipeline",
    )
    parser.add_argument(
        "--tulip-root",
        type=str,
        required=True,
        help="Root directory of the TULIP dataset",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default=None,
        help="Directory for generated CSVs (default: <tulip-root>/pipeline_csvs)",
    )
    parser.add_argument(
        "--cameras",
        type=str,
        default=None,
        help=(
            "Comma-separated camera numbers to include (e.g. '1,2,3'). "
            "Default: all cameras."
        ),
    )
    return parser


def main() -> None:
    parser = _build_parser()
    args = parser.parse_args()

    tulip_root = Path(args.tulip_root)
    if not tulip_root.is_dir():
        print(f"ERROR: TULIP root not found: {tulip_root}", file=sys.stderr)
        sys.exit(1)

    output_dir = Path(args.output_dir) if args.output_dir else tulip_root / "pipeline_csvs"

    cameras = None
    if args.cameras:
        cameras = [int(c.strip()) for c in args.cameras.split(",")]

    print(f"TULIP dataset preparation")
    print(f"  Root:   {tulip_root}")
    print(f"  Output: {output_dir}")
    if cameras:
        print(f"  Cameras: {cameras}")
    print()

    paths = generate_csvs(tulip_root, output_dir, cameras=cameras)

    print(f"\nDone. To run the pipeline on TULIP data:")
    print(f"  python scripts/run_pipeline.py --config configs/config_tulip.yaml")


if __name__ == "__main__":
    main()
